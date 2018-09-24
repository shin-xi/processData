import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('./example/example.xlsx')

print(wb.sheetnames, len(wb.sheetnames))

# print(wb['Sheet1'])

# print(wb.active)

sheet = wb['Sheet1']
# print(sheet['A1'])
# print(sheet['A1'].value, sheet['B1'].value, sheet['C1'].value)
# print('Row ' + str(sheet['A1'].row) + ', Column ' + sheet['A1'].column + ' is ' + sheet['A1'].value)
# print('Cell ' + sheet['A1'].coordinate + ' is ' + sheet['A1'].value)

# print(sheet.cell(row=1, column=2))
# print(sheet.cell(row=1, column=2).value)

# for i in range(1, 8, 2):
#     print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

print(get_column_letter(sheet.max_column))

print(column_index_from_string('A'))
print(column_index_from_string('AA'))

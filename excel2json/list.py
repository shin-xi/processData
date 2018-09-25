import openpyxl
import pprint
import json


def toList(file,sheet,target):
    wb = openpyxl.load_workbook(file)
    sheet = wb['Sheet1']
    row = sheet.max_row
    col = sheet.max_column
    listData = []
    for i in range(2, row + 1):
        item = {}
        for j in range(1, col + 1):
            item[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            listData.append(item)
    
    pprint.pprint(listData)
    list = open(target, 'w')
    list.write(json.dumps(listData))  


# wb = openpyxl.load_workbook('./list/list.xlsx')
# sheet = wb['Sheet1']
# pprint.pprint(toList(sheet))

# list1 = open('./list/list1.json', 'w')
# list1.write(json.dumps(toList(sheet)))
toList('./list/list.xlsx','Sheet1','./list/list1.json')